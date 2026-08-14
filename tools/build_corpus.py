#!/usr/bin/env python3
"""
Builds the deterministic test corpus for the Cause List Matcher.

Every planted case carries an ID (P01..P40) and an expected verdict, so the
golden file is generated from the same source of truth as the PDF. No manual
transcription, therefore no drift between fixture and expectation.
"""
import json, os, textwrap
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import openpyxl

OUT = "corpus"
os.makedirs(f"{OUT}/fixtures", exist_ok=True)

# ------------------------------------------------------------------
# THE FIRM  (mirrors the real chambers found in CAUSE_14082026.pdf)
# ------------------------------------------------------------------
FIRM = [
    dict(name="E. Ganesh",          enrol="MS/1234/2005", role="Senior"),
    dict(name="E. Srikanth",        enrol="MS/2255/2011", role="Junior"),
    dict(name="D. Lokeshwaran",     enrol="MS/3311/2014", role="Junior"),
    dict(name="T. Thamarai Selvan", enrol="MS/4088/2016", role="Junior"),
    dict(name="M. Krishnamurthy",   enrol="MS/0912/2001", role="Senior"),
    dict(name="S. Lakshmi Narayanan",enrol="MS/5521/2019", role="Junior"),
]

# ------------------------------------------------------------------
# PLANTED ITEMS.  spelling = what the registry actually printed.
# expect: auto | review | weak | none        why: the trap being tested
# ------------------------------------------------------------------
P = []
def plant(pid, court, hall, item, case, ptnr, resp, counsel, expect, why,
          firm_adv=None, resp_counsel=None, coram="A.B.JUSTICE.J", ltype="MOTION LIST"):
    P.append(dict(id=pid, court=court, hall=hall, item=item, case=case,
                  petitioner=ptnr, respondent=resp, counsel=counsel,
                  resp_counsel=resp_counsel or [], expect=expect, why=why,
                  firm_adv=firm_adv, coram=coram, ltype=ltype))

MHC = "HIGH COURT OF JUDICATURE AT MADRAS"

# --- A. exact and trivially-formatted matches -----------------------
plant("P01", MHC, "01", 1,  "WA/2025/2026",      "THE ASSISTANT COMMISSIONER", "MAJOR SKG RAMASWAMY AND 2 OTHERS",
      ["E.GANESH","E.SRIKANTH","D.LOKESHWARAN","T.THAMARAI SELVAN"], "auto",
      "exact spelling, four firm advocates stacked (cluster signal)", "E. Ganesh")
plant("P02", MHC, "01", 2,  "WP/14523/2025",     "R.KALAIVANI", "THE STATE AND 3 OTHERS",
      ["M/S.E.GANESH","R.PRABAKAR"], "auto", "M/S. chambers prefix must be stripped", "E. Ganesh")
plant("P03", MHC, "02", 3,  "CRP/414/2023",      "P.RENUKA AND 2 OTHERS", "SHAHUL HAMEED AND 12 OTHERS",
      ["K.BALU","M.R.ELAVARASAN"], "auto", "firm counsel below divider, FOR R13",
      "E. Ganesh", resp_counsel=["M/S.E.GANESH","R.PRABAKAR","E.SRIKANTH","FOR R13"])
plant("P04", MHC, "02", 4,  "CS/813/2016",       "DR.MOHAMMED HUSSAIN MUZAFER", "AMTUL BATHOOL AND 2 OTHERS",
      ["M/S.T.S.R.VENKATARAMANA","E.GANESH"], "auto", "second name in a chambers block", "E. Ganesh")
plant("P05", MHC, "03", 5,  "WMP/11451/2026",    "THE SECRETARY", "M RAJA",
      ["GANESH E."], "auto", "initial moved to the end", "E. Ganesh")

# --- B. Tamil transliteration variants ------------------------------
plant("P06", MHC, "03", 6,  "WP/9087/2026",      "S.ARUMUGAM", "THE DISTRICT COLLECTOR",
      ["M.KRISHNAMOORTHI"], "auto", "Krishnamurthy -> Krishnamoorthi (OO/EE fold)", "M. Krishnamurthy")
plant("P07", MHC, "04", 7,  "WP/9088/2026",      "K.VASANTHA", "THE COMMISSIONER",
      ["M.KIRUSHNAMURTHY"], "review", "epenthetic vowel inserted after K", "M. Krishnamurthy")
plant("P08", MHC, "04", 8,  "OSA/44/2024",       "TAMILARASI", "THE BANK AND ANOTHER",
      ["S.LAXMINARAYANAN"], "review", "Lakshmi Narayanan -> Laxmi (KSH->X) + joined", "S. Lakshmi Narayanan")
plant("P09", MHC, "05", 9,  "WP/9090/2026",      "M.SELVI", "THE STATE",
      ["T.THAMARAISELVAN"], "auto", "Thamarai Selvan joined into one token", "T. Thamarai Selvan")
plant("P10", MHC, "05", 10, "WP/9091/2026",      "A.PALANI", "THE SECRETARY",
      ["T.DHAMARAI SELVAN"], "review", "TH -> DH transliteration drift", "T. Thamarai Selvan")
plant("P11", MHC, "06", 11, "CMA/210/2025",      "N.RAJALAKSHMI", "THE MANAGER",
      ["D.LOKESWARAN"], "auto", "Lokeshwaran -> Lokeswaran (SH->S)", "D. Lokeshwaran")
plant("P12", MHC, "06", 12, "WP/9093/2026",      "G.MURUGAN", "THE TAHSILDAR",
      ["E.SREEKANTH"], "auto", "Srikanth -> Sreekanth (EE fold)", "E. Srikanth")

# --- C. initials: omission, typo, flip, transposition ---------------
plant("P13", MHC, "07", 13, "WP/9094/2026",      "P.ANBU", "THE STATE",
      ["GANESH"], "review", "initial omitted entirely - cannot disambiguate", "E. Ganesh")
plant("P14", MHC, "07", 14, "WP/9095/2026",      "V.MEENA", "THE COMMISSIONER",
      ["F.GANESH"], "review", "E->F lookalike typo in initial", "E. Ganesh")
plant("P15", MHC, "08", 15, "WP/9096/2026",      "S.BHARATHI", "THE STATE",
      ["R.GANESH"], "review", "E->R keyboard-adjacent initial", "E. Ganesh")
plant("P16", MHC, "08", 16, "WP/9097/2026",      "K.SARAVANAN", "THE DIRECTOR",
      ["E.S.GANESH"], "review", "extra initial inserted", "E. Ganesh")
plant("P17", MHC, "09", 17, "WP/9098/2026",      "R.DEVI", "THE SECRETARY",
      ["SELVAN T.THAMARAI"], "review", "name parts reordered", "T. Thamarai Selvan")

# --- D. DECOYS: must NOT surface in the confirm list -----------------
plant("D01", MHC, "09", 18, "WP/8001/2026",      "A.KUMAR", "THE STATE",
      ["J.GANESH"], "weak", "different advocate, J not near E", None)
plant("D02", MHC, "10", 19, "WP/8002/2026",      "B.RAJA", "THE STATE",
      ["M.GANESH"], "weak", "different advocate, M not near E", None)
plant("D03", MHC, "10", 20, "WP/8003/2026",      "C.VIJAY", "THE STATE",
      ["R.GANESHKUMAR"], "weak", "longer surname - Ganeshkumar != Ganesh", None)
plant("D04", MHC, "11", 21, "WP/8004/2026",      "D.MALA", "THE STATE",
      ["S.GANESH BABU"], "weak", "compound surname - different person", None)
plant("D05", MHC, "11", 22, "WP/8005/2026",      "E.RANI", "THE STATE",
      ["S.VIGNESH"], "none", "vowel-strip trap: VIGNESH vs GANESH must not match", None)
plant("D06", MHC, "12", 23, "WP/8006/2026",      "F.ARUN", "THE STATE",
      ["N.SIVAKUMAR"], "none", "common name, unrelated to the firm", None)
plant("D07", MHC, "12", 24, "WP/8007/2026",      "G.PRIYA", "THE STATE",
      ["K.SHIVAKUMAR"], "none", "common name variant, still unrelated", None)
plant("D08", MHC, "13", 25, "WP/8008/2026",      "H.BABU", "THE STATE",
      ["E.SRIKANTHAN"], "review", "near-miss on a firm junior - genuinely ambiguous", "E. Srikanth")
plant("D09", MHC, "13", 26, "WP/8009/2026",      "I.LATHA", "THE STATE",
      ["M.KRISHNAN"], "none", "Krishnan != Krishnamurthy", None)
plant("D10", MHC, "14", 27, "WP/8010/2026",      "J.SELVI", "THE STATE",
      ["T.SELVAN"], "none", "Selvan alone != Thamarai Selvan", None)

# --- E. case-number-driven matches (name absent or unreadable) ------
plant("P18", MHC, "14", 28, "CC/212/2026",       "E.GANESAN AND ANOTHER", "THE MANAGING DIRECTOR, M/S KLM ROYAL DUTCH AIRLINES",
      ["PARTY IN PERSON"], "auto", "matched on case number from the firm's register, not the name", None)
plant("P19", MHC, "15", 29, "OS/88/2024",        "S.RAMESH", "THE BANK",
      ["ADVOCATE NAME ILLEGIBLE"], "auto", "case number match rescues an unreadable counsel line", None)

# --- F. tribunal / district formats (different layout) --------------
plant("P20", "CONSUMER DISPUTES REDRESSAL COMMISSION, CHENNAI (SOUTH)", "1", 4,
      "CC/213/2026", "E.GANESH", "M/S KLM ROYAL DUTCH AIRLINES",
      ["COMPLAINANT IN PERSON"], "auto",
      "tribunal layout; party name is also the firm advocate's name", None,
      coram="PRESIDENT & MEMBER", ltype="DAILY BOARD")
plant("P21", "PRINCIPAL DISTRICT COURT, CHENGALPATTU", "2", 12,
      "OS/455/2022", "V.MURUGESAN", "THE VILLAGE ADMINISTRATIVE OFFICER",
      ["E.GANESH","T.THAMARAI SELVAN"], "auto",
      "district court layout, two firm advocates", "E. Ganesh",
      coram="PRL. DISTRICT JUDGE", ltype="REGULAR LIST")

# ------------------------------------------------------------------
# RENDER THE CAUSE LIST PDF  (mirrors Madras HC column geometry)
# ------------------------------------------------------------------
def render_pdf(path, items, date_str="FRIDAY 14 AUGUST 2026"):
    c = canvas.Canvas(path, pagesize=A4)
    W, H = A4
    LM, TM = 14*mm, H - 18*mm
    LEAD = 4.4*mm
    COL_ITEM, COL_CASE, COL_PARTY, COL_ADV = LM, LM+9*mm, LM+42*mm, LM+108*mm

    groups = {}
    for it in items:
        groups.setdefault((it["court"], it["hall"], it["coram"], it["ltype"]), []).append(it)

    page_no = 0
    for (court, hall, coram, ltype), rows in groups.items():
        page_no += 1
        y = TM
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(W/2, y, court); y -= LEAD*1.4
        c.setFont("Helvetica", 8.5)
        c.drawCentredString(W/2, y, date_str); y -= LEAD*1.2
        c.drawCentredString(W/2, y, f"CORAM : THE HON'BLE {coram}"); y -= LEAD*1.2
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(W/2, y, f"COURT NO. {hall}"); y -= LEAD*1.8

        c.setFont("Courier", 7.6)
        for it in rows:
            y0 = y
            c.drawString(COL_ITEM,  y, str(it["item"]))
            c.drawString(COL_CASE,  y, it["case"])
            c.drawString(COL_PARTY, y, it["petitioner"][:34])
            if it["counsel"]:
                c.drawString(COL_ADV, y, it["counsel"][0][:40])
            y -= LEAD
            c.drawString(COL_PARTY, y, "VS")
            if len(it["counsel"]) > 1:
                c.drawString(COL_ADV, y, it["counsel"][1][:40])
            y -= LEAD
            resp_wrapped = textwrap.wrap(it["respondent"], 34) or [""]
            extra = it["counsel"][2:]
            n = max(len(resp_wrapped), len(extra))
            for i in range(n):
                if i < len(resp_wrapped): c.drawString(COL_PARTY, y, resp_wrapped[i])
                if i < len(extra):        c.drawString(COL_ADV,   y, extra[i][:40])
                y -= LEAD
            c.drawString(COL_ADV, y, "-"*18); y -= LEAD
            for rc in it["resp_counsel"]:
                c.drawString(COL_ADV, y, rc[:40]); y -= LEAD
            y -= LEAD*0.6
            if y < 30*mm:
                c.setFont("Helvetica", 6.5)
                c.drawString(LM, 12*mm, f"{page_no}/1   TEST (14/08/2026) - {ltype}")
                c.showPage(); c.setFont("Courier", 7.6); y = TM
        c.setFont("Helvetica", 6.5)
        c.drawString(LM, 12*mm, f"{page_no}/1   TEST (14/08/2026) - {ltype}")
        c.showPage()
    c.save()

render_pdf(f"{OUT}/fixtures/causelist-synthetic-14082026.pdf", P)

# a second, deliberately messy file: no divider, ALL CAPS runs, odd spacing
messy = [dict(p, counsel=[x.replace(".", ". ") for x in p["counsel"]]) for p in P[:8]]
render_pdf(f"{OUT}/fixtures/causelist-messy-14082026.pdf", messy)

# ------------------------------------------------------------------
# ADVOCATE FIXTURES  (csv, xlsx, txt)
# ------------------------------------------------------------------
with open(f"{OUT}/fixtures/advocates.csv", "w", encoding="utf-8") as f:
    f.write("Advocate Name,Enrolment No,Role\n")
    for a in FIRM:
        f.write(f'{a["name"]},{a["enrol"]},{a["role"]}\n')

with open(f"{OUT}/fixtures/advocates.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(a["name"] for a in FIRM) + "\n")

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Advocates"
ws.append(["Advocate Name", "Enrolment No", "Role"])
for a in FIRM: ws.append([a["name"], a["enrol"], a["role"]])
wb.save(f"{OUT}/fixtures/advocates.xlsx")

# messy header variant - tests column auto-detection
wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2.title = "Sheet1"
ws2.append(["S.No", "NAME OF THE ADVOCATE", "Bar Enrolment", "Remarks"])
for i, a in enumerate(FIRM, 1): ws2.append([i, a["name"].upper(), a["enrol"], ""])
wb2.save(f"{OUT}/fixtures/advocates-messy-headers.xlsx")

# ------------------------------------------------------------------
# CASE REGISTER FIXTURE  (schema taken from the firm's Inventory.pdf)
# ------------------------------------------------------------------
CASE_COLS = ["DiaryNo","CaseType","CaseNo","Year","Court","CauseTitle","CounselFor",
             "PartyName","Mobile","Reference","Status","Date","Fees","Remarks",
             "AttendedBy","NextDate","NextStage","StatusRemark","CNR"]
CASES = [
 ["B208","CC","212","2026","Consumer Court, Chennai",
  "E.Ganesh & anr -Vs- The Managing Director, M/s KLM Royal Dutch Airlines","Complainant",
  "E.Ganesh","9840910000","Advocate Ganesh","Filed","04/02/2026","50% award amount",
  "Online filed","E. Ganesh","14/08/2026","Arguments","Counter awaited","TNCH01-000212-2026"],
 ["B209","CC","213","2026","Consumer Court, Chennai",
  "E.Ganesh -Vs- M/s KLM Royal Dutch Airlines","Complainant","E.Ganesh","9840910000",
  "Advocate Ganesh","Filed","04/02/2026","50% award amount","","E. Ganesh","14/08/2026",
  "Evidence","","TNCH01-000213-2026"],
 ["A128","WA","2025","2026","Madras High Court",
  "The Assistant Commissioner -Vs- Major SKG Ramaswamy and 2 others","Appellant",
  "The Assistant Commissioner","9840911111","Referred","Pending","10/03/2026","",
  "Counter to be filed","E. Ganesh","14/08/2026","Counter","Opposite side sought time","TNMC01-002025-2026"],
 ["A129","WP","14523","2025","Madras High Court",
  "R.Kalaivani -Vs- The State and 3 others","Petitioner","R.Kalaivani","9840912222",
  "Walk-in","Pending","22/07/2025","","","E. Srikanth","14/08/2026","Arguments","","TNMC01-014523-2025"],
 ["C234","CRP","414","2023","Madras High Court",
  "P.Renuka and 2 others -Vs- Shahul Hameed and 12 others","Respondent 13","Shahul Hameed",
  "9840913333","Referred","Pending","05/05/2023","","Appearing for R13","E. Ganesh",
  "14/08/2026","Arguments","","TNMC01-000414-2023"],
 ["C011","CS","813","2016","Madras High Court",
  "Dr.Mohammed Hussain Muzafer -Vs- Amtul Bathool and 2 others","Plaintiff",
  "Dr.Mohammed Hussain Muzafer","9840914444","Referred","Pending","11/11/2016","","",
  "E. Ganesh","14/08/2026","Evidence","","TNMC01-000813-2016"],
 ["B045","OS","88","2024","Principal District Court, Chengalpattu",
  "S.Ramesh -Vs- The Bank","Plaintiff","S.Ramesh","9840915555","Walk-in","Pending",
  "01/02/2024","","","D. Lokeshwaran","14/08/2026","Trial","","TNCG02-000088-2024"],
 ["B046","OS","455","2022","Principal District Court, Chengalpattu",
  "V.Murugesan -Vs- The Village Administrative Officer","Plaintiff","V.Murugesan",
  "9840916666","Referred","Pending","09/09/2022","","","T. Thamarai Selvan","14/08/2026",
  "Arguments","","TNCG02-000455-2022"],
 ["A188","WP","7777","2026","Madras High Court",
  "K.Anbarasan -Vs- The State","Petitioner","K.Anbarasan","9840917777","Walk-in","Pending",
  "01/06/2026","","NOT LISTED TODAY - control row","E. Ganesh","20/08/2026","Counter","","TNMC01-007777-2026"],
]
with open(f"{OUT}/fixtures/cases.csv", "w", encoding="utf-8") as f:
    f.write(",".join(CASE_COLS) + "\n")
    for r in CASES:
        f.write(",".join('"'+str(x).replace('"','""')+'"' for x in r) + "\n")

wb3 = openpyxl.Workbook(); ws3 = wb3.active; ws3.title = "Cases"
ws3.append(CASE_COLS)
for r in CASES: ws3.append(r)
wb3.save(f"{OUT}/fixtures/cases.xlsx")

# ------------------------------------------------------------------
# GOLDEN EXPECTATIONS
# ------------------------------------------------------------------
golden = dict(
    generated_from="build_corpus.py",
    causelist="fixtures/causelist-synthetic-14082026.pdf",
    date="2026-08-14",
    firm=[a["name"] for a in FIRM],
    total_planted=len(P),
    expected_counts=dict(
        auto=sum(1 for p in P if p["expect"] == "auto"),
        review=sum(1 for p in P if p["expect"] == "review"),
        weak=sum(1 for p in P if p["expect"] == "weak"),
        none=sum(1 for p in P if p["expect"] == "none"),
    ),
    items=[dict(id=p["id"], item=p["item"], case=p["case"], court=p["court"],
                hall=p["hall"], printed_counsel=p["counsel"] + p["resp_counsel"],
                expect=p["expect"], firm_advocate=p["firm_adv"], why=p["why"])
           for p in P],
)
with open(f"{OUT}/fixtures/golden.json", "w", encoding="utf-8") as f:
    json.dump(golden, f, indent=2)

print(f"planted items      : {len(P)}")
for k, v in golden["expected_counts"].items():
    print(f"  expect {k:<7}: {v}")
print("\nfiles written:")
for root, _, fs in os.walk(OUT):
    for fn in sorted(fs):
        fp = os.path.join(root, fn)
        print(f"  {fp:<58} {os.path.getsize(fp):>8,} bytes")
